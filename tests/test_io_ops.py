from __future__ import annotations

import io
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

sys.path.append(str(Path(__file__).resolve().parents[1] / "src"))

from patent_app.exporter import TEMPLATE_OUTPUT_COLUMNS, build_xlsx_bytes
from patent_app.io_ops import (
    INTERNAL_PUBLICATION_URL_COLUMN,
    canonicalize_dataframe,
    load_dataframe,
    resolve_column_mapping,
)


def test_excel_header_is_detected_from_second_row() -> None:
    source = pd.DataFrame(
        [
            ["Patent Export, 2026-05-13 02:09:46 +0000", None, None, None],
            ["DWPI アクセッション番号", "公報番号", "出願日", "出願番号"],
            ["1991325232", "JP8500001A", "1991-04-11", "JP1991508026A"],
        ]
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        source.to_excel(writer, index=False, header=False)

    df = load_dataframe("sample.xlsx", buf.getvalue())

    assert list(df.columns)[:4] == [
        "DWPI アクセッション番号",
        "公報番号",
        "出願日",
        "出願番号",
    ]
    assert df.iloc[0]["公報番号"] == "JP8500001A"


def test_load_dataframe_extracts_publication_url_from_hyperlink_cell() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Patent Export, 2026-05-13 02:09:46 +0000", None, None])
    ws.append(["公報番号", "出願番号", "出願日"])
    ws.append(["JP8500001A", "JP1991508026A", "1991-04-11"])
    ws["A3"].hyperlink = "https://example.com/patent/JP8500001A"

    buf = io.BytesIO()
    wb.save(buf)

    df = load_dataframe("sample.xlsx", buf.getvalue())

    assert "publication_url" not in df.columns
    assert INTERNAL_PUBLICATION_URL_COLUMN in df.columns
    assert df.iloc[0][INTERNAL_PUBLICATION_URL_COLUMN] == "https://example.com/patent/JP8500001A"


def test_country_code_is_generated_from_publication_number() -> None:
    source = pd.DataFrame(
        {
            "公報番号": ["EP1234567A1"],
            "国コード": ["JP"],
        }
    )

    mapping = resolve_column_mapping(list(source.columns))
    out = canonicalize_dataframe(source, mapping)

    assert out.iloc[0]["country_code"] == "EP"


def test_canonicalize_dataframe_keeps_non_canonical_columns() -> None:
    source = pd.DataFrame(
        {
            "公報番号": ["JP20240001A1"],
            "明細 (英語)": ["Detailed specification"],
        }
    )

    mapping = resolve_column_mapping(list(source.columns))
    out = canonicalize_dataframe(source, mapping)

    assert "明細 (英語)" in out.columns
    assert out.iloc[0]["明細 (英語)"] == "Detailed specification"


def test_canonicalize_dataframe_normalizes_non_breaking_spaces() -> None:
    source = pd.DataFrame(
        {
            "公報番号": ["JP20240001A1"],
            "請求項（英語）": ["1. A\u00A0method\u00A0comprising."],
        }
    )

    mapping = resolve_column_mapping(list(source.columns))
    out = canonicalize_dataframe(source, mapping)

    assert out.iloc[0]["請求項（英語）"] == "1. A method comprising."


def test_publication_date_is_mapped_from_koho_hakkobi() -> None:
    source = pd.DataFrame(
        {
            "公報番号": ["JP8500001A"],
            "公報発行日": ["1996-01-09"],
        }
    )

    mapping = resolve_column_mapping(list(source.columns))
    out = canonicalize_dataframe(source, mapping)

    assert out.iloc[0]["publication_date"] == date(1996, 1, 9)


def test_legal_status_is_binary_valid_invalid() -> None:
    source = pd.DataFrame(
        {
            "公報番号": ["JP8500001A", "JP6000001B2"],
            "法的状況": ["active", "失効"],
        }
    )

    mapping = resolve_column_mapping(list(source.columns))
    out = canonicalize_dataframe(source, mapping)

    # 元の表記をそのまま保持する
    assert out.iloc[0]["legal_status"] == "active"
    assert out.iloc[1]["legal_status"] == "失効"


def test_dead_status_is_normalized_to_invalid() -> None:
    source = pd.DataFrame(
        {
            "公報番号": ["JP2008293226A"],
            "法的状況": ["Dead"],
        }
    )

    mapping = resolve_column_mapping(list(source.columns))
    out = canonicalize_dataframe(source, mapping)

    # 元の表記「Dead」をそのまま保持する
    assert out.iloc[0]["legal_status"] == "Dead"


def test_mukou_yukou_column_is_mapped_to_legal_status() -> None:
    source = pd.DataFrame(
        {
            "公報番号": ["JP5502366A"],
            "無効/有効": ["Dead"],
        }
    )

    mapping = resolve_column_mapping(list(source.columns))
    out = canonicalize_dataframe(source, mapping)

    # 元の表記「Dead」をそのまま保持する
    assert out.iloc[0]["legal_status"] == "Dead"


def test_kind_is_kind_code_suffix() -> None:
    source = pd.DataFrame(
        {
            "公報番号": ["JP8500001A", "JP6000001B2"],
        }
    )

    mapping = resolve_column_mapping(list(source.columns))
    out = canonicalize_dataframe(source, mapping)

    assert out.iloc[0]["kind"] == "A"
    assert out.iloc[1]["kind"] == "B2"


def test_kind_code_routes_registration_from_publication_column() -> None:
    source = pd.DataFrame(
        {
            "公報番号": ["JP6000001B2"],
            "出願番号": ["JP1984258790A"],
        }
    )

    mapping = resolve_column_mapping(list(source.columns))
    out = canonicalize_dataframe(source, mapping)

    assert out.iloc[0]["publication_number"] == ""
    assert out.iloc[0]["registration_number"] == "JP6000001B2"


def test_kind_code_routes_publication_from_registration_column() -> None:
    source = pd.DataFrame(
        {
            "登録番号": ["JP8500001A"],
            "出願番号": ["JP1991508026A"],
        }
    )

    mapping = resolve_column_mapping(list(source.columns))
    out = canonicalize_dataframe(source, mapping)

    assert out.iloc[0]["publication_number"] == "JP8500001A"
    assert out.iloc[0]["registration_number"] == ""


def test_kr_publication_application_number_07_is_converted_to_70_after_1998() -> None:
    source = pd.DataFrame(
        {
            "公報番号": ["KR2001123456A"],
            "出願番号": ["KR200107123456"],
            "出願日": ["2001-10-01"],
        }
    )

    mapping = resolve_column_mapping(list(source.columns))
    out = canonicalize_dataframe(source, mapping)

    assert out.iloc[0]["application_number"] == "KR200170123456"


def test_kr_registration_application_number_70_is_converted_to_07_before_1998() -> None:
    source = pd.DataFrame(
        {
            "公報番号": ["KR1995123456B"],
            "出願番号": ["KR199570123456"],
            "出願日": ["1995-03-20"],
        }
    )

    mapping = resolve_column_mapping(list(source.columns))
    out = canonicalize_dataframe(source, mapping)

    assert out.iloc[0]["application_number"] == "KR199507123456"


def test_template_export_uses_screener_column_schema() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                "registration_number": "JP7654321B2",
                "publication_date": pd.Timestamp("2024-02-01"),
                "application_number": "JP20230001",
                "application_date": pd.Timestamp("2023-01-15"),
                "priority_number": "PRIO-1",
                "priority_date": "2022-12-01 | 2023-01-10",
                "優先権情報": "PRIO-1(2022-12-01)",
                "DWPI ファミリーメンバー": "JP20240001A1|US20240001A1",
                "dwpi_family_members_status": "JP20240001A1 Alive|US20240001A1 Dead",
                "accession_number": "ACC-001",
                "language_of_publication": "JA",
                "title_english": "Sample title",
                "title_dwpi": "DWPI title",
                "assignee_applicant": "Applicant A",
                "assignee_dwpi": "DWPI Applicant",
                "assignee_standardized": "Standardized Applicant",
                "請求項数": "18",
                "IPC - 最新": "A01B 1/00",
                "US クラス": "435/6",
                "CPC - 最新": "A01B1/00",
                "抄録（英語）": "English abstract",
                "抄録 - DWPI 優位性": "DWPI advantage",
                "抄録 - DWPI 新規性": "DWPI novelty",
                "抄録 - DWPI 用途": "DWPI use",
                "請求項 (英語)": "English claims",
                "INPADOC ファミリーメンバー": "JP20240001A1|US20240001A1",
                "五庁有効ファミリ": "JP20240001A1",
                "五庁失効ファミリ": "US20240001A1",
                "その他ファミリ": "WO2024000001A1",
                "legal_status": "Alive",
                "country_code": "JP",
                "source_file": "sample.xlsx",
            }
        ]
    )

    template_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SearchData"
    ws.append(["placeholder"])
    wb.save(template_buffer)

    output_bytes = build_xlsx_bytes(selected, template_bytes=template_buffer.getvalue())
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    assert headers == TEMPLATE_OUTPUT_COLUMNS

    values = [cell.value for cell in result_ws[2]]
    row = dict(zip(headers, values))
    assert row["NUID"] == 1
    assert row["DATE"] is None
    assert row["NUMBER"] == "JP20240001A1"
    assert row["公報番号"] == "JP20240001A1"
    assert row["公報言語"] == "JA"
    assert row["タイトル (英語)"] == "Sample title"
    assert row["タイトル - DWPI"] == "DWPI title"
    assert row["譲受人/出願人"] == "Applicant A"
    assert row["譲受人 - DWPI"] == "DWPI Applicant"
    assert row["譲受人 - 標準化"] == "Standardized Applicant"
    assert row["請求項数"] == "18"
    assert row["IPC - 最新"] == "A01B 1/00"
    assert row["US クラス"] == "435/6"
    assert row["CPC - 最新"] == "A01B1/00"
    assert row["抄録 (英語)"] == "English abstract"
    assert row["抄録 - DWPI 優位性"] == "DWPI advantage"
    assert row["抄録 - DWPI 新規性"] == "DWPI novelty"
    assert row["抄録 - DWPI 用途"] == "DWPI use"
    assert row["請求項 (英語)"] == "English claims"
    assert row["DWPI アクセッション番号"] == "ACC-001"
    assert row["公報発行日"].date() == date(2024, 2, 1)
    assert row["出願番号"] == "JP20230001"
    assert row["出願日"].date() == date(2023, 1, 15)
    assert row["優先権主張番号"] == "PRIO-1"
    assert row["優先権主張日"] == "2022-12-01 | 2023-01-10"
    assert row["優先権情報"] == "PRIO-1(2022-12-01)"
    assert row["DWPI ファミリーメンバー"] == "JP20240001A1|US20240001A1"
    assert row["DWPI ファミリーメンバー 有効/無効"] == "JP20240001A1 Alive|US20240001A1 Dead"
    assert row["INPADOC ファミリーメンバー"] == "JP20240001A1|US20240001A1"
    assert row["FileName"] == "sample.xlsx"
    assert row["公開番号"] == "JP20240001A1"
    assert row["登録番号"] == "JP7654321B2"
    assert row["五庁有効ファミリ"] == "JP20240001A1"
    assert row["五庁失効ファミリ"] == "US20240001A1"
    assert row["その他ファミリ"] == "WO2024000001A1"
    assert row["無効/有効"] == "Alive"
    assert row["国名コード"] == "JP"
    assert row["PDF コピー"] is None


def test_template_export_falls_back_to_applicant_rights_column() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                "registration_number": "",
                "publication_date": pd.Timestamp("2024-02-01"),
                "application_number": "JP20230001",
                "application_date": pd.Timestamp("2023-01-15"),
                "出願人/権利者": "Applicant From Pipeline",
            }
        ]
    )

    template_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SearchData"
    ws.append(["placeholder"])
    wb.save(template_buffer)

    output_bytes = build_xlsx_bytes(selected, template_bytes=template_buffer.getvalue())
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    values = [cell.value for cell in result_ws[2]]
    row = dict(zip(headers, values))
    assert row["譲受人/出願人"] == "Applicant From Pipeline"


def test_template_export_fills_title_columns_from_japanese_named_sources() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                "タイトル（英語）": "Title From Pipeline",
                "タイトル - DWPI": "DWPI From Pipeline",
            }
        ]
    )

    template_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SearchData"
    ws.append(["placeholder"])
    wb.save(template_buffer)

    output_bytes = build_xlsx_bytes(selected, template_bytes=template_buffer.getvalue())
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    values = [cell.value for cell in result_ws[2]]
    row = dict(zip(headers, values))
    assert row["タイトル (英語)"] == "Title From Pipeline"
    assert row["タイトル - DWPI"] == "DWPI From Pipeline"


def test_template_export_fills_claims_column_from_fullwidth_header() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                "請求項（英語）": "Claims From Pipeline",
            }
        ]
    )

    template_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SearchData"
    ws.append(["placeholder"])
    wb.save(template_buffer)

    output_bytes = build_xlsx_bytes(selected, template_bytes=template_buffer.getvalue())
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    values = [cell.value for cell in result_ws[2]]
    row = dict(zip(headers, values))
    assert row["請求項 (英語)"] == "Claims From Pipeline"


def test_template_export_appends_passthrough_columns_after_fixed_schema() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                "source_file": "sample.xlsx",
                "明細 (英語)": "Detailed specification",
                "Custom Score": "A+",
            }
        ]
    )

    template_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SearchData"
    ws.append(["placeholder"])
    wb.save(template_buffer)

    output_bytes = build_xlsx_bytes(selected, template_bytes=template_buffer.getvalue())
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    assert headers[: len(TEMPLATE_OUTPUT_COLUMNS)] == TEMPLATE_OUTPUT_COLUMNS
    assert headers[-2:] == ["明細 (英語)", "Custom Score"]

    values = [cell.value for cell in result_ws[2]]
    row = dict(zip(headers, values))
    assert row["明細 (英語)"] == "Detailed specification"
    assert row["Custom Score"] == "A+"


def test_template_export_extracts_independent_claim_numbers_and_hides_raw_column() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                "独立請求項": "請求項1. AAA\n請求項10. BBB\n請求項12. CCC",
                "source_file": "sample.xlsx",
            }
        ]
    )

    template_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SearchData"
    ws.append(["placeholder"])
    wb.save(template_buffer)

    output_bytes = build_xlsx_bytes(selected, template_bytes=template_buffer.getvalue())
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    assert "独立請求項" not in headers

    values = [cell.value for cell in result_ws[2]]
    row = dict(zip(headers, values))
    assert row["独立請求項番号"] == "1,10,12"


def test_template_export_prefers_existing_independent_claim_numbers() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                "独立請求項番号": "3,7",
                "独立請求項": "請求項1. AAA\n請求項10. BBB",
            }
        ]
    )

    template_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SearchData"
    ws.append(["placeholder"])
    wb.save(template_buffer)

    output_bytes = build_xlsx_bytes(selected, template_bytes=template_buffer.getvalue())
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    values = [cell.value for cell in result_ws[2]]
    row = dict(zip(headers, values))
    assert row["独立請求項番号"] == "3,7"


def test_template_export_independent_claim_numbers_fallback_per_row() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                "独立請求項": "請求項2. AAA",
                "請求項（英語）": "1. A method.\n2. The method of claim 1.",
            },
            {
                "selected_patent_number": "JP20240002A1",
                "publication_number": "JP20240002A1",
                "独立請求項": "",
                "請求項（英語）": "1. A device.\n2. The device of claim 1.",
            },
        ]
    )

    template_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SearchData"
    ws.append(["placeholder"])
    wb.save(template_buffer)

    output_bytes = build_xlsx_bytes(selected, template_bytes=template_buffer.getvalue())
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    row1 = dict(zip(headers, [cell.value for cell in result_ws[2]]))
    row2 = dict(zip(headers, [cell.value for cell in result_ws[3]]))

    assert row1["独立請求項番号"] == "2"
    assert row2["独立請求項番号"] == "1"


def test_template_export_extracts_independent_claim_numbers_from_claims_english() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                "請求項（英語）": (
                    "1. A composition comprising A and B.\n"
                    "2. The composition of claim 1, wherein A is X.\n"
                    "3. A method for preparing the composition."
                ),
            }
        ]
    )

    template_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SearchData"
    ws.append(["placeholder"])
    wb.save(template_buffer)

    output_bytes = build_xlsx_bytes(selected, template_bytes=template_buffer.getvalue())
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    values = [cell.value for cell in result_ws[2]]
    row = dict(zip(headers, values))
    assert row["独立請求項番号"] == "1,3"


def test_template_export_excludes_first_claim_when_canceled_in_claims_english() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                "請求項 (英語)": (
                    "1. (canceled)\n"
                    "2. The apparatus of claim 1, wherein the member is fixed.\n"
                    "3. An apparatus comprising a sensor and controller."
                ),
            }
        ]
    )

    template_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SearchData"
    ws.append(["placeholder"])
    wb.save(template_buffer)

    output_bytes = build_xlsx_bytes(selected, template_bytes=template_buffer.getvalue())
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    values = [cell.value for cell in result_ws[2]]
    row = dict(zip(headers, values))
    assert row["独立請求項番号"] == "3"


def test_template_export_parses_pipe_prefixed_claims_and_excludes_dependent_claims() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                "請求項（英語）": (
                    "CLAIMS:\n"
                    " | 1. A computer-implemented method comprising steps A and B.\n"
                    " | 2. The method of claim 1, wherein step A uses model X.\n"
                    " | 3. The method of claim 1, wherein step B uses model Y."
                ),
            }
        ]
    )

    template_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SearchData"
    ws.append(["placeholder"])
    wb.save(template_buffer)

    output_bytes = build_xlsx_bytes(selected, template_bytes=template_buffer.getvalue())
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    values = [cell.value for cell in result_ws[2]]
    row = dict(zip(headers, values))
    assert row["独立請求項番号"] == "1"


def test_non_template_export_keeps_passthrough_columns() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                "country_code": "JP",
                "明細 (英語)": "Detailed specification",
            }
        ]
    )

    output_bytes = build_xlsx_bytes(selected)
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    assert "明細 (英語)" in headers

    values = [cell.value for cell in result_ws[2]]
    row = dict(zip(headers, values))
    assert row["明細 (英語)"] == "Detailed specification"


def test_template_export_embeds_hyperlink_in_publication_number_cells() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                INTERNAL_PUBLICATION_URL_COLUMN: "https://example.com/patent/JP20240001A1",
            }
        ]
    )

    template_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SearchData"
    ws.append(["placeholder"])
    wb.save(template_buffer)

    output_bytes = build_xlsx_bytes(selected, template_bytes=template_buffer.getvalue())
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    number_col = headers.index("NUMBER") + 1
    koho_col = headers.index("公報番号") + 1

    number_cell = result_ws.cell(row=2, column=number_col)
    koho_cell = result_ws.cell(row=2, column=koho_col)

    assert number_cell.value == "JP20240001A1"
    assert koho_cell.value == "JP20240001A1"
    assert number_cell.hyperlink is None
    assert koho_cell.hyperlink is not None
    assert koho_cell.hyperlink.target == "https://example.com/patent/JP20240001A1"


def test_non_template_export_embeds_hyperlink_in_publication_number_cell() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                INTERNAL_PUBLICATION_URL_COLUMN: "https://example.com/p/JP20240001A1",
            }
        ]
    )

    output_bytes = build_xlsx_bytes(selected)
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    koho_col = headers.index("公報番号") + 1
    koho_cell = result_ws.cell(row=2, column=koho_col)

    assert koho_cell.value == "JP20240001A1"
    assert koho_cell.hyperlink is not None
    assert koho_cell.hyperlink.target == "https://example.com/p/JP20240001A1"


def test_non_template_export_skips_invalid_publication_url() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1",
                INTERNAL_PUBLICATION_URL_COLUMN: "www.example.com/no-scheme",
            }
        ]
    )

    output_bytes = build_xlsx_bytes(selected)
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    koho_col = headers.index("公報番号") + 1
    koho_cell = result_ws.cell(row=2, column=koho_col)

    assert koho_cell.value == "JP20240001A1"
    assert koho_cell.hyperlink is None


def test_non_template_export_does_not_extract_url_from_publication_number_text() -> None:
    selected = pd.DataFrame(
        [
            {
                "selected_patent_number": "JP20240001A1",
                "publication_number": "JP20240001A1 https://example.com/q/JP20240001A1)",
            }
        ]
    )

    output_bytes = build_xlsx_bytes(selected)
    result_wb = load_workbook(io.BytesIO(output_bytes))
    result_ws = result_wb["SearchData"]

    headers = [cell.value for cell in result_ws[1]]
    koho_col = headers.index("公報番号") + 1
    koho_cell = result_ws.cell(row=2, column=koho_col)

    assert koho_cell.value == "JP20240001A1"
    assert koho_cell.hyperlink is None
