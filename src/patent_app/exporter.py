from __future__ import annotations

import io
import re
import posixpath
import zipfile
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

OUTPUT_COLUMN_RENAME: dict[str, str] = {
    "country_code": "国名コード",
    "accession_number": "ファミリ番号",
    "selected_patent_number": "公報番号",
    "publication_number": "公開番号",
    "registration_number": "登録番号",
    "publication_date": "公報発行日",
    "application_number": "出願番号",
    "application_date": "出願日",
    "legal_status": "無効/有効",
    "source_file": "ソースファイル",
    "language_of_publication": "公報言語",
}

TEMPLATE_OUTPUT_COLUMNS = [
    "NUID",
    "DATE",
    "SEARCHER",
    "CATEGORY",
    "NUMBER",
    "JUDGMENT",
    "CODE",
    "RANK",
    "OTHER",
    "MEMO1",
    "MEMO2",
    "公報番号",
    "PDF コピー",
    "請求項数",
    "公報言語",
    "タイトル (英語)",
    "タイトル - DWPI",
    "譲受人/出願人",
    "譲受人 - DWPI",
    "譲受人 - 標準化",
    "IPC - 最新",
    "US クラス",
    "CPC - 最新",
    "フロントページ イメージ",
    "フロントページ図",
    "抄録 (英語)",
    "抄録 - DWPI 優位性",
    "抄録 - DWPI 新規性",
    "抄録 - DWPI 用途",
    "請求項 (英語)",
    "DWPI アクセッション番号",
    "公報発行日",
    "出願番号",
    "出願日",
    "優先権主張番号",
    "優先権主張日",
    "優先権情報",
    "DWPI ファミリーメンバー",
    "DWPI ファミリーメンバー 有効/無効",
    "INPADOC ファミリーメンバー",
    "独立請求項番号",
    "FileName",
    "公開番号",
    "登録番号",
    "五庁有効ファミリ",
    "五庁失効ファミリ",
    "その他ファミリ",
    "無効/有効",
    "国名コード",
    "単一効 (EP)",
    "発明者",
]

TEMPLATE_OUTPUT_SOURCE_MAP: dict[str, str | list[str]] = {
    "NUMBER": "selected_patent_number",
    "公報番号": "selected_patent_number",
    "請求項数": ["請求項数", "claim_count", "number_of_claims", "claims_count"],
    "公報言語": "language_of_publication",
    "タイトル (英語)": ["title_english", "タイトル (英語)", "タイトル（英語）"],
    "タイトル - DWPI": ["title_dwpi", "タイトル - DWPI"],
    "譲受人/出願人": ["assignee_applicant", "出願人/権利者"],
    "譲受人 - DWPI": ["譲受人 - DWPI", "assignee_dwpi"],
    "譲受人 - 標準化": ["譲受人 - 標準化", "assignee_standardized"],
    "IPC - 最新": ["IPC - 最新", "ipc_latest", "latest_ipc", "ipc_classification"],
    "US クラス": ["US クラス", "us_class", "us_classification", "uspc_class"],
    "CPC - 最新": ["CPC - 最新", "cpc_latest", "latest_cpc", "cpc_classification"],
    "抄録 (英語)": ["抄録 (英語)", "抄録（英語）", "abstract_english", "abstract (english)", "abstract_eng"],
    "抄録 - DWPI 優位性": [
        "抄録 - DWPI 優位性",
        "abstract_dwpi_advantage",
        "dwpi_abstract_advantage",
        "abstract-dwpi-advantage",
    ],
    "抄録 - DWPI 新規性": [
        "抄録 - DWPI 新規性",
        "abstract_dwpi_novelty",
        "dwpi_abstract_novelty",
        "abstract-dwpi-novelty",
    ],
    "抄録 - DWPI 用途": ["抄録 - DWPI 用途", "abstract_dwpi_use", "dwpi_abstract_use", "abstract-dwpi-use"],
    "請求項 (英語)": ["請求項 (英語)", "請求項（英語）", "claims_english", "claims (english)", "claim_text"],
    "DWPI アクセッション番号": "accession_number",
    "公報発行日": "publication_date",
    "出願番号": "application_number",
    "出願日": "application_date",
    "優先権主張番号": ["優先権主張番号", "priority_number"],
    "優先権主張日": ["優先権主張日", "priority_date"],
    "優先権情報": ["優先権情報", "priority_number"],
    "DWPI ファミリーメンバー": ["DWPI ファミリーメンバー", "dwpi_family_members"],
    "DWPI ファミリーメンバー 有効/無効": ["DWPI ファミリーメンバー 有効/無効", "dwpi_family_members_status"],
    "INPADOC ファミリーメンバー": ["INPADOC ファミリーメンバー", "inpadoc_family_members", "inpadoc_family"],
    "独立請求項番号": [
        "独立請求項番号",
        "独立請求項",
        "請求項 (英語)",
        "請求項（英語）",
        "claims_english",
        "claims (english)",
        "claim_text",
    ],
    "FileName": "source_file",
    "公開番号": "publication_number",
    "登録番号": "registration_number",
    "五庁有効ファミリ": "五庁有効ファミリ",
    "五庁失効ファミリ": "五庁失効ファミリ",
    "その他ファミリ": "その他ファミリ",
    "無効/有効": ["無効/有効", "legal_status"],
    "国名コード": ["国名コード", "country_code"],
    "単一効 (EP)": ["単一効 (EP)", "単一効(EP)"],
    "発明者": ["発明者", "inventor", "inventors"],
}

TEMPLATE_SUPPRESSED_PASSTHROUGH_COLUMNS = {
    "独立請求項",
}

TEMPLATE_BLANK_COLUMNS = {
    "DATE",
    "SEARCHER",
    "CATEGORY",
    "JUDGMENT",
    "CODE",
    "RANK",
    "OTHER",
    "MEMO1",
    "MEMO2",
    "PDF コピー",
    "請求項数",
    "IPC - 最新",
    "US クラス",
    "CPC - 最新",
    "フロントページ イメージ",
    "フロントページ図",
    "抄録 (英語)",
    "抄録 - DWPI 優位性",
    "抄録 - DWPI 新規性",
    "抄録 - DWPI 用途",
    "請求項 (英語)",
    "INPADOC ファミリーメンバー",
    "独立請求項番号",
}


def build_xlsx_bytes(
    selected_df: pd.DataFrame,
    no_acc_df: pd.DataFrame | None = None,
    template_bytes: bytes | None = None,
    keep_vba: bool = False,
) -> bytes:
    if template_bytes:
        wb = _load_template_workbook(template_bytes, keep_vba)
    else:
        wb = Workbook()

    results_ws = _get_or_create_sheet(wb, "SearchData")
    export_df = _build_template_output_dataframe(selected_df) if template_bytes else selected_df.rename(
        columns=OUTPUT_COLUMN_RENAME
    )
    _write_dataframe(results_ws, export_df)

    if "NoAcc" in wb.sheetnames:
        del wb["NoAcc"]

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _get_or_create_sheet(wb: Workbook, title: str):
    if title in wb.sheetnames:
        return wb[title]
    return wb.create_sheet(title)


def _write_dataframe(ws, df: pd.DataFrame) -> None:
    ws.delete_rows(1, ws.max_row)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)


def _build_template_output_dataframe(selected_df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame(index=selected_df.index)
    row_count = len(selected_df)
    consumed_source_columns: set[str] = set()

    for column in TEMPLATE_OUTPUT_COLUMNS:
        source = TEMPLATE_OUTPUT_SOURCE_MAP.get(column)
        if source is None:
            out[column] = [None] * row_count
            continue

        source_candidates = [source] if isinstance(source, str) else source
        if column == "独立請求項番号":
            present_candidates = [candidate for candidate in source_candidates if candidate in selected_df.columns]
            if not present_candidates:
                out[column] = [None] * row_count
                continue
            consumed_source_columns.update(present_candidates)
            out[column] = _build_independent_claim_number_series(selected_df, present_candidates)
            continue

        source_column = next((candidate for candidate in source_candidates if candidate in selected_df.columns), None)
        if source_column is None:
            out[column] = [None] * row_count
            continue
        consumed_source_columns.add(source_column)

        values = selected_df[source_column]
        if column == "NUMBER" or column == "公報番号":
            out[column] = values.fillna("").astype(str)
        elif column in {"公報発行日", "出願日"}:
            out[column] = pd.to_datetime(values, errors="coerce").dt.date
        elif column == "優先権主張日":
            out[column] = values.map(_normalize_priority_date_text)
        else:
            out[column] = values.map(_normalize_export_text)

    for column in selected_df.columns:
        if column in TEMPLATE_OUTPUT_COLUMNS:
            continue
        if column in consumed_source_columns:
            continue
        if column in TEMPLATE_SUPPRESSED_PASSTHROUGH_COLUMNS:
            continue
        out[column] = selected_df[column].map(_normalize_export_text)

    out["NUID"] = range(1, row_count + 1)
    return out


def _normalize_export_text(value: object) -> object:
    if pd.isna(value):
        return None
    text = str(value).strip()
    return text if text else None


def _normalize_priority_date_text(value: object) -> object:
    if pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()

    text = str(value).strip()
    if not text:
        return None

    parsed = pd.to_datetime(text, errors="coerce")
    if not pd.isna(parsed):
        return parsed.date().isoformat()

    return text


def _normalize_independent_claim_numbers(value: object, source_column: str) -> object:
    if pd.isna(value):
        return None

    text = str(value).strip()
    if not text:
        return None

    if source_column == "独立請求項番号":
        return _normalize_export_text(text)

    if source_column in {"請求項 (英語)", "請求項（英語）", "claims_english", "claims (english)", "claim_text"}:
        return _extract_independent_claim_numbers_from_english_claims(text)

    return _extract_claim_numbers_from_text(text)


def _build_independent_claim_number_series(selected_df: pd.DataFrame, source_candidates: list[str]) -> pd.Series:
    values: list[object] = []
    for _, row in selected_df.iterrows():
        picked: object = None
        for candidate in source_candidates:
            raw = row.get(candidate)
            normalized = _normalize_independent_claim_numbers(raw, candidate)
            if normalized is None:
                continue
            if str(normalized).strip() == "":
                continue
            picked = normalized
            break
        values.append(picked)
    return pd.Series(values, index=selected_df.index)


def _extract_claim_numbers_from_text(text: str) -> object:
    normalized = text.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
    parts = re.split(r"[\n\r\t,、;；|]+", normalized)

    numbers: list[int] = []
    seen: set[int] = set()
    for part in parts:
        token = str(part).strip()
        if not token:
            continue
        match = re.match(r"^(?:【?請求項\s*)?[\(\[]?(\d+)", token)
        if match:
            claim_no = int(match.group(1))
            if claim_no not in seen:
                seen.add(claim_no)
                numbers.append(claim_no)

    if not numbers:
        for found in re.findall(r"請求項\s*(\d+)", normalized):
            claim_no = int(found)
            if claim_no not in seen:
                seen.add(claim_no)
                numbers.append(claim_no)

    if not numbers:
        return _normalize_export_text(text)

    numbers.sort()
    return ",".join(str(num) for num in numbers)


def _extract_independent_claim_numbers_from_english_claims(text: str) -> object:
    normalized = text.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
    pairs = _parse_english_claim_pairs(normalized)
    if not pairs:
        return _extract_claim_numbers_from_text(text)

    independent_numbers: list[int] = []
    seen: set[int] = set()
    for idx, (claim_no, claim_body) in enumerate(pairs):
        body_lower = claim_body.lower()
        is_canceled = "canceled" in body_lower or "cancelled" in body_lower
        contains_claim_word = bool(re.search(r"\bclaims?\b", body_lower))
        is_independent = (not is_canceled) and ((idx == 0) or (not contains_claim_word))
        if is_independent and claim_no not in seen:
            seen.add(claim_no)
            independent_numbers.append(claim_no)

    if not independent_numbers:
        return None

    independent_numbers.sort()
    return ",".join(str(num) for num in independent_numbers)


def _parse_english_claim_pairs(text: str) -> list[tuple[int, str]]:
    start_pattern = re.compile(
        r"(?:^|\n)\s*(?:[|¦]\s*)?(?:claims?\s*)?(\d+)\s*[\.)\]:]\s*",
        flags=re.IGNORECASE,
    )
    matches = list(start_pattern.finditer(text))
    if not matches:
        return []

    pairs: list[tuple[int, str]] = []
    for idx, match in enumerate(matches):
        claim_no = int(match.group(1))
        body_start = match.end()
        body_end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
        claim_body = text[body_start:body_end].strip()
        pairs.append((claim_no, claim_body))

    return pairs


def _load_template_workbook(template_bytes: bytes, keep_vba: bool) -> Workbook:
    try:
        return load_workbook(io.BytesIO(template_bytes), keep_vba=keep_vba)
    except KeyError as exc:
        missing_item = _extract_missing_archive_item(str(exc))
        if not missing_item:
            raise

        repaired_bytes = _remove_broken_relationships(template_bytes, missing_item)
        try:
            return load_workbook(io.BytesIO(repaired_bytes), keep_vba=keep_vba)
        except Exception:
            # Some templates contain chained drawing/image relationship corruption.
            # Fall back to a clean workbook so export can still complete.
            return Workbook()


def _extract_missing_archive_item(error_message: str) -> str:
    match = re.search(r"'([^']+)'", error_message)
    if not match:
        return ""
    return match.group(1).lstrip("/")


def _remove_broken_relationships(template_bytes: bytes, missing_item: str) -> bytes:
    missing_norm = missing_item.lstrip("/")
    in_buffer = io.BytesIO(template_bytes)
    out_buffer = io.BytesIO()

    with zipfile.ZipFile(in_buffer, "r") as zin, zipfile.ZipFile(out_buffer, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)

            if not info.filename.endswith(".rels"):
                zout.writestr(info, data)
                continue

            fixed = _strip_invalid_relationship_entries(data, info.filename, missing_norm)
            zout.writestr(info, fixed)

    out_buffer.seek(0)
    return out_buffer.read()


def _strip_invalid_relationship_entries(xml_bytes: bytes, rels_path: str, missing_item: str) -> bytes:
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return xml_bytes

    namespace = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    base_dir = _resolve_relationship_base_dir(rels_path)
    removed = False

    for rel in list(root.findall(f"{namespace}Relationship")):
        if rel.get("TargetMode") == "External":
            continue

        target = str(rel.get("Target", "") or "").strip()
        if not target:
            continue

        resolved = _resolve_target_path(base_dir, target)
        if resolved == missing_item or target.upper().endswith("/NULL") or target.upper() == "NULL":
            root.remove(rel)
            removed = True

    if not removed:
        return xml_bytes

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _resolve_relationship_base_dir(rels_path: str) -> str:
    parent = posixpath.dirname(rels_path)
    if parent.endswith("/_rels"):
        parent = parent[: -len("/_rels")]

    filename = posixpath.basename(rels_path)
    if filename.endswith(".rels"):
        source_part = filename[: -len(".rels")]
    else:
        source_part = filename

    source_path = posixpath.join(parent, source_part)
    return posixpath.dirname(source_path)


def _resolve_target_path(base_dir: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(base_dir, target)).lstrip("/")
